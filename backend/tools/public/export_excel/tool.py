"""生成 Excel 并上传 OSS，供前端下载。"""

from __future__ import annotations

import io
import json
import logging
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from langchain_core.tools import tool

from backend.common.tool_outcome import format_tool_user_message

logger = logging.getLogger(__name__)

TOOL_NAME = "export_excel"

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_UNSAFE_NAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_MAX_ROWS = 5000
_MAX_COLS = 40
_MAX_TITLE_LEN = 64
_MAX_SHEET_LEN = 31


def _safe_filename(name: str) -> str:
    name = (name or "").strip() or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    name = _UNSAFE_NAME.sub("_", name)
    name = name.replace("..", "_").strip(" .")
    if not name:
        name = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if not name.lower().endswith(".xlsx"):
        if "." in name:
            name = name.rsplit(".", 1)[0] or name
        name = f"{name}.xlsx"
    return name


def _safe_sheet_name(name: str) -> str:
    text = (name or "").strip() or "Sheet1"
    text = re.sub(r'[:\\/?*\[\]]', "_", text)
    if len(text) > _MAX_SHEET_LEN:
        text = text[:_MAX_SHEET_LEN]
    return text or "Sheet1"


def _parse_jsonish(raw: Any) -> Any:
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return raw
    return raw


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _normalize_column_defs(raw: Any) -> list[dict[str, str]]:
    """将 columns 规范为 [{key, title}, ...]。"""
    data = _parse_jsonish(raw)
    if data is None:
        return []
    if not isinstance(data, list):
        raise ValueError("columns 须为数组，例如 [\"姓名\",\"金额\"] 或 [{\"key\":\"amount\",\"title\":\"金额\"}]。")

    out: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for item in data:
        if len(out) >= _MAX_COLS:
            break
        if isinstance(item, str):
            title = item.strip()
            key = title
        elif isinstance(item, dict):
            key = str(item.get("key") or item.get("field") or item.get("name") or "").strip()
            title = str(item.get("title") or item.get("label") or key).strip()
            if not key:
                key = title
        else:
            continue
        if not title and not key:
            continue
        if not title:
            title = key
        if not key:
            key = title
        title = title[:_MAX_TITLE_LEN]
        # 重复 key 时加后缀，避免行取值互相覆盖
        base = key
        n = 2
        while key in seen_keys:
            key = f"{base}_{n}"
            n += 1
        seen_keys.add(key)
        out.append({"key": key, "title": title})
    return out


def _normalize_rows(raw: Any) -> list[Any]:
    data = _parse_jsonish(raw)
    if data is None:
        return []
    if not isinstance(data, list):
        raise ValueError("rows 须为对象数组或二维数组。")
    if len(data) > _MAX_ROWS:
        raise ValueError(f"行数超过上限（最多 {_MAX_ROWS} 行）。")
    return data


def _infer_columns_from_rows(rows: list[Any]) -> list[dict[str, str]]:
    keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        for k in row.keys():
            key = str(k).strip()
            if not key or key in seen:
                continue
            seen.add(key)
            keys.append(key)
            if len(keys) >= _MAX_COLS:
                return [{"key": k, "title": k[:_MAX_TITLE_LEN]} for k in keys]
    return [{"key": k, "title": k[:_MAX_TITLE_LEN]} for k in keys]


def _row_values(row: Any, columns: list[dict[str, str]]) -> list[Any]:
    if isinstance(row, dict):
        return [_cell_value(row.get(col["key"], "")) for col in columns]
    if isinstance(row, (list, tuple)):
        vals = list(row)
        out: list[Any] = []
        for i, _ in enumerate(columns):
            out.append(_cell_value(vals[i] if i < len(vals) else ""))
        return out
    # 单值行：只填第一列
    out = [""] * len(columns)
    if columns:
        out[0] = _cell_value(row)
    return out


# LLM 常见误用：竖表「字段|内容」——自动转成「表头横排、一行一条记录」
_KV_FIELD_LABELS = frozenset(
    {"字段", "field", "key", "名称", "项目", "属性", "栏目"}
)
_KV_VALUE_LABELS = frozenset(
    {"内容", "值", "value", "content", "数值", "数据"}
)


def _label_norm(text: str) -> str:
    return (text or "").strip().lower()


def _is_kv_label_pair(col_a: dict[str, str], col_b: dict[str, str]) -> bool:
    labels_a = {_label_norm(col_a.get("key", "")), _label_norm(col_a.get("title", ""))}
    labels_b = {_label_norm(col_b.get("key", "")), _label_norm(col_b.get("title", ""))}
    field_set = {_label_norm(x) for x in _KV_FIELD_LABELS}
    value_set = {_label_norm(x) for x in _KV_VALUE_LABELS}
    a_is_field = bool(labels_a & field_set)
    b_is_value = bool(labels_b & value_set)
    a_is_value = bool(labels_a & value_set)
    b_is_field = bool(labels_b & field_set)
    return (a_is_field and b_is_value) or (a_is_value and b_is_field)


def _widen_kv_layout(
    row_list: list[Any],
    col_defs: list[dict[str, str]],
) -> tuple[list[Any], list[dict[str, str]]]:
    """若是「字段/内容」竖表，转为宽表：各字段名作表头，整单一行。"""
    if len(col_defs) != 2 or len(row_list) < 2:
        return row_list, col_defs
    if not _is_kv_label_pair(col_defs[0], col_defs[1]):
        return row_list, col_defs

    labels_0 = {
        _label_norm(col_defs[0].get("key", "")),
        _label_norm(col_defs[0].get("title", "")),
    }
    field_set = {_label_norm(x) for x in _KV_FIELD_LABELS}
    if labels_0 & field_set:
        field_col, value_col = col_defs[0], col_defs[1]
    else:
        field_col, value_col = col_defs[1], col_defs[0]

    field_key = field_col["key"]
    value_key = value_col["key"]
    # 也兼容二维数组 [["付款人户名","刘健"], ...]
    wide: dict[str, Any] = {}
    order: list[str] = []
    for row in row_list:
        if isinstance(row, dict):
            name = str(row.get(field_key) or row.get(field_col["title"]) or "").strip()
            val = row.get(value_key, row.get(value_col["title"], ""))
        elif isinstance(row, (list, tuple)) and len(row) >= 2:
            name = str(row[0] or "").strip()
            val = row[1]
        else:
            continue
        if not name or _label_norm(name) in {_label_norm(x) for x in _KV_FIELD_LABELS}:
            continue
        if name not in wide:
            order.append(name)
        wide[name] = val

    if len(order) < 2:
        return row_list, col_defs

    new_cols = [{"key": k, "title": k[:_MAX_TITLE_LEN]} for k in order[:_MAX_COLS]]
    logger.info(
        "export_excel: widened field/content layout → %d columns, 1 row",
        len(new_cols),
    )
    return [wide], new_cols


def build_excel_bytes(
    rows: Any,
    columns: Any = None,
    sheet_name: str = "Sheet1",
) -> bytes:
    """将结构化表格写成 xlsx bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    row_list = _normalize_rows(rows)
    col_defs = _normalize_column_defs(columns)
    if not col_defs:
        if row_list and isinstance(row_list[0], (list, tuple)):
            width = min(max(len(r) for r in row_list if isinstance(r, (list, tuple))), _MAX_COLS)
            if width <= 0:
                raise ValueError("表格为空，请提供 columns 或含字段的 rows。")
            col_defs = [
                {"key": f"col_{i + 1}", "title": f"列{i + 1}"} for i in range(width)
            ]
        else:
            col_defs = _infer_columns_from_rows(row_list)
    if not col_defs:
        raise ValueError("表格为空，请提供 columns 或含字段的 rows。")

    row_list, col_defs = _widen_kv_layout(row_list, col_defs)

    if len(col_defs) > _MAX_COLS:
        col_defs = col_defs[:_MAX_COLS]

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)
    header_font = Font(bold=True)
    for c_idx, col in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col["title"])
        cell.font = header_font

    for r_idx, row in enumerate(row_list, start=2):
        for c_idx, value in enumerate(_row_values(row, col_defs), start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 简单列宽：按表头与抽样内容
    for c_idx, col in enumerate(col_defs, start=1):
        sample_len = len(col["title"])
        for row in row_list[:50]:
            vals = _row_values(row, col_defs)
            sample_len = max(sample_len, len(str(vals[c_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(
            max(sample_len + 2, 8), 40
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel_to_oss(
    rows: Any,
    columns: Any = None,
    filename: str = "",
    sheet_name: str = "Sheet1",
) -> dict:
    """生成 Excel 并上传 OSS，返回前端可用的下载元信息。"""
    from backend.common.oss import build_file_url, put_object, sign_get_url

    data = build_excel_bytes(rows, columns=columns, sheet_name=sheet_name)
    safe_name = _safe_filename(filename)
    object_key = f"chat-exports/{uuid.uuid4().hex[:12]}_{safe_name}"
    put_object(object_key, data, _XLSX_MIME)
    permanent = build_file_url(object_key)
    download_url = sign_get_url(object_key, expires=3600 * 24)
    logger.info("Exported excel to OSS: %s (%d bytes)", object_key, len(data))
    return {
        "name": safe_name,
        "url": download_url,
        "permanent_url": permanent,
        "object_key": object_key,
        "mime_type": _XLSX_MIME,
        "file_size": len(data),
    }


def format_export_excel_result(meta: dict) -> str:
    name = meta.get("name") or "export.xlsx"
    return format_tool_user_message(
        f"表格「{name}」已准备好，界面会显示下载卡片。",
        ask="请用一两句友好中文告知用户可以点击下方卡片下载到本机；",
    )


@tool(TOOL_NAME)
def export_excel(
    rows: list,
    columns: list | None = None,
    filename: str = "",
    sheet_name: str = "Sheet1",
) -> str:
    """导出 Excel（.xlsx）。表头为列、每行一条记录；勿传散文。
    表格数据已就绪或用户短确认导出时调用；勿为导出再检索。

    Args:
        rows: 行数据（对象数组，或配合 columns 的二维数组）
        columns: 表头；字符串数组或 [{key,title}]；对象行可省略
        filename: 文件名，空=自动命名
        sheet_name: 工作表名，默认 Sheet1
    """
    try:
        meta = export_excel_to_oss(
            rows, columns=columns, filename=filename, sheet_name=sheet_name
        )
    except ValueError as e:
        from backend.common.errors import tool_user_error

        return tool_user_error("导出", e)
    except Exception as e:
        logger.exception("export_excel failed")
        from backend.common.errors import tool_user_error

        return tool_user_error("导出", e)
    return format_export_excel_result(meta)
